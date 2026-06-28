from openpyxl import load_workbook

from models.cell import Cell


class LayoutLoader:

    def __init__(self, path):
        self.path = path

    def load(self):

        wb = load_workbook(self.path, data_only=True)
        ws = wb.active

        cells = []

        for row in ws.iter_rows():

            for c in row:

                cells.append(
                    Cell(
                        id=c.coordinate,
                        row=c.row,
                        col=c.column,
                        value=c.value
                    )
                )

        return cells